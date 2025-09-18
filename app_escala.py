import streamlit as st
import pandas as pd
import random
import calendar
import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# Importa o banco de dados do outro arquivo
from banco_dados import igrejas

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Gerador de Escala", layout="centered")


class GeradorEscala:
    """
    Encapsula toda a lógica para gerar a escala de forma mais justa e organizada.
    """
    def __init__(self, mes, ano, dados_igreja):
        self.mes = mes
        self.ano = ano
        self.dados_igreja = dados_igreja
        
        # Extrai os dados da igreja para uso na classe
        self.homens_sede = dados_igreja["homens_sede"]
        self.mulheres_sede = dados_igreja["mulheres_sede"]
        self.homens_nao_sede = dados_igreja["homens_nao_sede"]
        self.mulheres_nao_sede = dados_igreja["mulheres_nao_sede"]
        
        # Estruturas de controle
        self.escala_anterior = {}
        self.contadores = {nome: 0 for nome in self.homens_sede + self.mulheres_sede + self.homens_nao_sede + self.mulheres_nao_sede}
        
        # Funções
        self.funcoes_homens_single = ["Lateral do Púlpito", "Porta Direita"]
        self.funcao_mulher = "Porta Esquerda"
        self.funcao_homens_dupla = "Porta Central"

    def _obter_dias_de_culto(self):
        dias = []
        cal = calendar.monthcalendar(self.ano, self.mes)
        sabados_do_mes = []
        for semana in cal:
            for dia_da_semana, dia_do_mes in enumerate(semana):
                if dia_do_mes == 0: continue
                
                if dia_da_semana in [2, 4, 6]: # Qua, Sex, Dom
                    dias.append(datetime.date(self.ano, self.mes, dia_do_mes))
                if dia_da_semana == 5: # Sábado
                    sabados_do_mes.append(datetime.date(self.ano, self.mes, dia_do_mes))

        if len(sabados_do_mes) >= 3:
            dias.append(sabados_do_mes[2]) # Adiciona o 3º Sábado
        
        return sorted(dias)

    def _escolher_obreiro(self, candidatos, funcao, ja_escalados):
        """Lógica aprimorada para escolher o próximo obreiro de forma justa."""
        # Regra 1: Não pode já estar escalado hoje
        pool = [p for p in candidatos if p not in ja_escalados]
        
        # Regra 2: Não pode ter servido nesta mesma função no último culto
        pessoas_na_funcao_anterior = self.escala_anterior.get(funcao, [])
        pool_sem_repeticao = [p for p in pool if p not in pessoas_na_funcao_anterior]
        
        # Se o filtro anti-repetição esvaziar a lista, relaxamos essa regra
        if not pool_sem_repeticao:
            pool_sem_repeticao = pool

        if not pool_sem_repeticao:
            return "N/D"

        # Regra 3 (Justiça): Prioriza quem trabalhou menos
        pool_sem_repeticao.sort(key=lambda nome: self.contadores.get(nome, 0))
        min_contagem = self.contadores.get(pool_sem_repeticao[0], 0)
        
        # Seleciona todos os candidatos com a menor contagem de turnos
        melhores_candidatos = [p for p in pool_sem_repeticao if self.contadores.get(p, 0) == min_contagem]
        
        # Escolhe aleatoriamente entre os melhores candidatos
        escolhido = random.choice(melhores_candidatos)
        
        # Atualiza o contador
        self.contadores[escolhido] += 1
        
        return escolhido

    def gerar(self):
        dias_de_culto = self._obter_dias_de_culto()
        dados_escala = []

        for data in dias_de_culto:
            escala_do_dia = {}
            ja_escalados_hoje = set()
            
            is_ceia = data.weekday() == 5
            is_quarta = data.weekday() == 2
            
            evento = ""
            if is_ceia: evento = "Santa Ceia do Senhor"
            elif is_quarta: evento = "Quarta-Feira"
            elif data.weekday() == 4: evento = "Sexta-Feira"
            else: evento = "Domingo Noite"

            if is_quarta or is_ceia:
                candidatos_homens = self.homens_sede + self.homens_nao_sede
                candidatas_mulheres = self.mulheres_sede + self.mulheres_nao_sede
            else:
                candidatos_homens = self.homens_sede.copy()
                candidatas_mulheres = self.mulheres_sede.copy()

            # --- Escala do Dia ---
            # Porta Central (2 Homens)
            homem1 = self._escolher_obreiro(candidatos_homens, self.funcao_homens_dupla, ja_escalados_hoje)
            ja_escalados_hoje.add(homem1)
            homem2 = self._escolher_obreiro(candidatos_homens, self.funcao_homens_dupla, ja_escalados_hoje)
            ja_escalados_hoje.add(homem2)
            escala_do_dia[self.funcao_homens_dupla] = [homem1, homem2]
            
            # Outras funções dos homens
            for funcao in self.funcoes_homens_single:
                homem = self._escolher_obreiro(candidatos_homens, funcao, ja_escalados_hoje)
                ja_escalados_hoje.add(homem)
                escala_do_dia[funcao] = [homem]

            # Função da mulher
            mulher = self._escolher_obreiro(candidatas_mulheres, self.funcao_mulher, ja_escalados_hoje)
            escala_do_dia[self.funcao_mulher] = [mulher]

            self.escala_anterior = escala_do_dia.copy()

            dados_escala.append({
                "Mês da escala": evento,
                "Data": data.strftime("%d/%m"),
                "Porta Central": f"{escala_do_dia[self.funcao_homens_dupla][0]} e {escala_do_dia[self.funcao_homens_dupla][1]}",
                "Porta Esquerda": escala_do_dia[self.funcao_mulher][0],
                "Lateral do Púlpito": escala_do_dia[self.funcoes_homens_single[0]][0],
                "Porta Direita": escala_do_dia[self.funcoes_homens_single[1]][0]
            })

        return pd.DataFrame(dados_escala)

def criar_excel_formatado(df, mes_nome, ano, logo_path):
    """Cria o arquivo .xlsx formatado em memória."""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Escala {mes_nome}"

    try:
        img = Image(logo_path)
        img.anchor = 'A1'
        img.height = 83 
        img.width = 94 
        ws.add_image(img)
    except FileNotFoundError:
        ws['A1'] = f"Logo não encontrada em {logo_path}"

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:F2')
    titulo_cell = ws['A1']
    titulo_cell.value = f"ESCALA DE OBREIROS - {mes_nome.upper()}/{ano}"
    titulo_cell.font = Font(name='Calibri', bold=True, size=26)
    titulo_cell.alignment = center_align
    ws.row_dimensions[1].height = 33
    ws.row_dimensions[2].height = 33

    colunas = list(df.columns)
    colunas[0] = f"Mês da escala ({mes_nome})"
    
    header_font = Font(name='Calibri', bold=True, color="000000", size=12)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    header_row_index = 4
    for col_idx, column_title in enumerate(colunas, 1):
        cell = ws.cell(row=header_row_index, column=col_idx, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for r_idx, row_data in enumerate(df.itertuples(index=False), 5):
        for c_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
            if c_idx == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            cell.border = thin_border
    
    last_row = ws.max_row
    atencao_textos = [
        "ATENÇÃO!!!",
        "Em caso de impossibilidade de atender a escala, o obreiro DEVE:",
        "1º Comunicar-se com outro obreiro que atenda sua escala",
        "2º Notificar ao pastor presidente quanto a ausência e troca"
    ]
    
    start_row_box = last_row + 2
    end_row_box = start_row_box + len(atencao_textos) - 1
    start_col_box = 1
    end_col_box = 3

    for i, texto in enumerate(atencao_textos):
        row_num = start_row_box + i
        ws.merge_cells(start_row=row_num, start_column=start_col_box, end_row=row_num, end_column=end_col_box)
        cell = ws.cell(row=row_num, column=start_col_box)
        cell.value = texto
        cell.alignment = center_align

        if i == 0:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(name='Calibri', bold=True)
            ws.row_dimensions[row_num].height = 22
        else:
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            if i == 1:
                cell.font = Font(name='Calibri', bold=True, color="FF3300")
            else:
                cell.font = Font(name='Calibri', bold=False)

    thick_side = Side(style="thick")
    for row in ws.iter_rows(min_row=start_row_box, max_row=end_row_box, min_col=start_col_box, max_col=end_col_box):
        for cell in row:
            existing_border = cell.border.copy()
            if cell.row == start_row_box: existing_border.top = thick_side
            if cell.row == end_row_box: existing_border.bottom = thick_side
            if cell.column == start_col_box: existing_border.left = thick_side
            if cell.column == end_col_box: existing_border.right = thick_side
            cell.border = existing_border

    ws.column_dimensions[get_column_letter(1)].width = 27
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 36
    ws.column_dimensions[get_column_letter(4)].width = 20
    ws.column_dimensions[get_column_letter(5)].width = 20
    ws.column_dimensions[get_column_letter(6)].width = 20

    wb.save(output)
    output.seek(0)
    return output


# --- INTERFACE DO USUÁRIO (UI) ---
st.title("👨‍👩‍👧‍👦 Gerador de Escala da Igreja")
st.markdown("Uma ferramenta para criar a escala de obreiros de forma justa e automática.")

with st.sidebar:
    st.header("🗓️ Configurações")

    # Seletor de Igreja
    igreja_selecionada_nome = st.selectbox(
        "Selecione a Igreja",
        options=list(igrejas.keys())
    )
    dados_igreja_selecionada = igrejas[igreja_selecionada_nome]

    
    meses_pt = {
        "Janeiro": 1, "Fevereiro": 2, "Março": 3, "Abril": 4, "Maio": 5, "Junho": 6,
        "Julho": 7, "Agosto": 8, "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12
    }
    
    mes_nome_selecionado = st.selectbox(
        "Selecione o Mês",
        options=list(meses_pt.keys()),
        index=datetime.date.today().month - 1
    )
    mes_selecionado = meses_pt[mes_nome_selecionado]
    
    ano_selecionado = st.number_input(
        "Digite o Ano",
        min_value=2024,
        max_value=2100,
        value=datetime.date.today().year
    )
    
    gerar_btn = st.button("Gerar Escala", type="primary", use_container_width=True)

if gerar_btn:
    with st.spinner("Gerando a escala... Por favor, aguarde."):
        # Instancia a classe e gera a escala
        gerador = GeradorEscala(mes_selecionado, ano_selecionado, dados_igreja_selecionada)
        df_escala = gerador.gerar()

    if df_escala is not None and not df_escala.empty:
        st.header("✅ Escala Gerada com Sucesso!")
        st.markdown("### Pré-visualização")
        st.dataframe(df_escala, use_container_width=True)

        excel_bytes = criar_excel_formatado(df_escala, mes_nome_selecionado, ano_selecionado, dados_igreja_selecionada['logo_path'])
        st.download_button(
            label="📥 Baixar Arquivo Excel (.xlsx)",
            data=excel_bytes,
            file_name=f"escala_{igreja_selecionada_nome.replace(' ', '_')}_{mes_nome_selecionado.lower()}_{ano_selecionado}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.warning("Não foi possível gerar a escala. Verifique os erros acima ou se há dias de culto no mês selecionado.")

st.markdown("---")
st.write("Desenvolvido para otimizar o trabalho do Reino.")

